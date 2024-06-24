import json
import openpyxl as xl
from openpyxl.styles.fonts import Font
from datetime import datetime

RAW_FACTOR = 2
TIME_FACTOR = 100


def excel_writer():
    # open workbook
    wb = xl.Workbook()
    etf_sheet = wb.active
    etf_sheet.title = "ETFs"

    etf_sheet['A1'] = 'Score Rank'
    etf_sheet['B1'] = 'Index'
    etf_sheet['C1'] = 'Provider'
    etf_sheet['D1'] = 'Ticker'
    etf_sheet['E1'] = 'Remaining Cap %'
    etf_sheet['F1'] = 'Remaining Buffer %'
    etf_sheet['G1'] = 'Downside Before Buffer %'
    etf_sheet['H1'] = 'Remaining Outcome Period'
    etf_sheet['I1'] = 'Starting Cap %'
    etf_sheet['J1'] = 'Cap Used %'
    etf_sheet['K1'] = 'Percentage Used %'
    etf_sheet['L1'] = 'Remaining Percentage %'
    etf_sheet['M1'] = 'One Day Potential Return %'
    etf_sheet['N1'] = 'One Month Potential Return %'
    etf_sheet['O1'] = 'One Year Potential Return %'
    etf_sheet['P1'] = 'Raw Value Sum * Mult'
    etf_sheet['Q1'] = 'Raw Value Mult'
    etf_sheet['Q2'] = RAW_FACTOR

    # get json
    with open('etf_data.json', 'r') as json_file:
        etf_data = json.load(json_file)

        current_cell = 2
        for provider, data in etf_data.items():
            for etf, values in data.items():
                # Score Rank
                etf_sheet.cell(current_cell, 2).value = current_cell - 1
                etf_sheet.cell(current_cell, 3).value = provider  # Provider
                etf_sheet.cell(current_cell, 4).value = etf  # Ticker
                etf_sheet.cell(
                    current_cell, 5).value = values['remaining_cap']  # Remaining Cap %
                etf_sheet.cell(
                    current_cell, 6).value = values['remaining_buffer']  # Remaining Buffer %
                # Downside Before Buffer %
                etf_sheet.cell(
                    current_cell, 7).value = values['downside_before_buffer']
                # Remaining Outcome Period (days)
                etf_sheet.cell(
                    current_cell, 8).value = values['remaining_outcome_period']
                etf_sheet.cell(
                    current_cell, 9).value = values['starting_cap']  # Starting Cap %
                etf_sheet.cell(
                    current_cell, 10).value = f'=I{current_cell} - E{current_cell}'  # Cap Used %
                # Percentage Used %
                etf_sheet.cell(
                    current_cell, 11).value = f'=J{current_cell}/I{current_cell}'
                # Remaining Percentage %
                etf_sheet.cell(current_cell, 12).value = f'=1-K{current_cell}'
                # One day potential return %
                etf_sheet.cell(
                    current_cell, 13).value = f'=E{current_cell}/H{current_cell}'
                # One month Potential Return %
                etf_sheet.cell(
                    current_cell, 14).value = f'=M{current_cell}*30.5'
                # One Year Potential Return %
                etf_sheet.cell(
                    current_cell, 15).value = f'=M{current_cell}*365'
                # Raw Sum * Mult
                etf_sheet.cell(
                    current_cell, 16).value = f'=(E{current_cell}+F{current_cell}+G{current_cell})*$Q$2'

                current_cell += 1

    # cap_used % = starting cap - remaining cap
    # percentage_used = cap_used/starting_cap
    # remaining_percentage = 100% - %used_up

    # one_day_potential_return = remaining_perecentage / remaining_days

    # write to Calculations Sheet

    calc_sheet = wb.create_sheet(title='Calc')

    calc_sheet['A1'] = 'Ticker'
    calc_sheet['B1'] = 'Risk/Reward'
    calc_sheet['C1'] = 'Period/Time Factor'
    calc_sheet['D1'] = 'Rank'
    calc_sheet['E1'] = 'Downside/Buffer'
    calc_sheet['F1'] = 'Score'

    calc_sheet['G1'] = 'Time Factor'
    calc_sheet['G2'] = TIME_FACTOR

    total_written = current_cell - 2

    for i in range(total_written):
        calc_sheet.cell(i + 2, 1).value = f'=ETFs!D{i+2}'  # Ticker
        calc_sheet.cell(
            i + 2, 2).value = f'=ETFs!E{i+2}/ETFs!G{i+2}'  # B column
        calc_sheet.cell(i + 2, 3).value = f'=$G$2/ETFs!H{i+2}'  # C Column
        calc_sheet.cell(i + 2, 4).value = f'=B{i+2}*C{i+2}'  # D Column
        calc_sheet.cell(
            i + 2, 5).value = f'=ETFs!G{i+2}/ETFs!F{i+2}'  # E column
        calc_sheet.cell(
            i + 2, 6).value = f'=D{i+2}+E{i+2}+ETFs!P{i+2}'  # F Column

    # perform ranking of the ETFS into a summary page
    with open('etf_data.json', 'r') as json_file:
        etf_data = json.load(json_file)

        summary_sheet = wb.create_sheet(title='Summary')

        calc_list = generate_calc_list(etf_data)

        # SHOWN BY OVERALL SCORE
        next_row_start = add_summary_rank(
            summary_sheet, 'score', calc_list, 1, 10)
        next_row_start = add_summary_rank(
            summary_sheet, 'risk/reward', calc_list, next_row_start, 10)
        next_row_start = add_summary_rank(
            summary_sheet, 'period/time_factor', calc_list, next_row_start, 10)
        next_row_start = add_summary_rank(
            summary_sheet, 'downside/buffer', calc_list, next_row_start, 10)

    # write score ranking equations in ETF sheet
    for i in range(total_written):
        formula = f'=_xlfn.RANK.EQ(Calc!F{i+2}, Calc!$F$2:Calc!$F${total_written+1})'
        etf_sheet.cell(row=i + 2, column=1, value=formula)

    filename = f'ETFReport_{datetime.now().strftime("%m-%d-%Y")}.xlsx'

    # resize all columns in all worksheets and format columns with percentages
    for sheet in wb.sheetnames:
        resize_columns(wb[sheet])
        format_percentages(wb[sheet])

    wb.save(filename)
    wb.close()

    return filename


def add_summary_rank(summary_sheet, summary_key, calc_list, start_row, num_to_display):
    # headers for this sort
    summary_sheet[f'A{start_row}'] = f'{summary_key}'
    start_row += 1

    summary_sheet[f'A{start_row}'] = 'Overall Rank'
    summary_sheet[f'B{start_row}'] = 'Ticker'
    summary_sheet[f'C{start_row}'] = 'Score'
    summary_sheet[f'D{start_row}'] = 'Risk/Reward'
    summary_sheet[f'E{start_row}'] = 'Period/Time Factor'
    summary_sheet[f'F{start_row}'] = "Downside/Buffer"

    summary_column_nums = {'score': 3, 'risk/reward': 4,
                           'period/time_factor': 5, 'downside/buffer': 6}
    start_row += 1

    # special font to signify what the summary is sorted by
    summary_column_font = Font(name='Calibri', b=True)

    sorted_calc_list = sort_calc_list(calc_list, sort_key=summary_key)
    for i in range(num_to_display):
        summary_sheet.cell(start_row, 1).value = i + 1  # rank for this sort
        summary_sheet.cell(start_row, 2).value = sorted_calc_list[i]['ticker']
        summary_sheet.cell(start_row, 3).value = sorted_calc_list[i]['score']
        summary_sheet.cell(
            start_row, 4).value = sorted_calc_list[i]['risk/reward']
        summary_sheet.cell(
            start_row, 5).value = sorted_calc_list[i]['period/time_factor']
        summary_sheet.cell(
            start_row, 6).value = sorted_calc_list[i]['downside/buffer']

        summary_sheet.cell(
            start_row, summary_column_nums[summary_key]).font = summary_column_font

        start_row += 1

    # give the next possible row a different summary could start on
    return start_row + 2


def generate_calc_list(etf_data):
    # list will store all etfs and their appropriate calculations
    calc_list = []

    for provider, etfs in etf_data.items():
        for ticker, etf_info in etfs.items():
            calc_dict = {}
            calc_dict['ticker'] = ticker
            calc_dict['raw_sum_mult'] = (
                etf_info['remaining_cap'] + etf_info['remaining_buffer'] + etf_info['downside_before_buffer']) * RAW_FACTOR
            calc_dict['risk/reward'] = etf_info['remaining_cap'] / \
                etf_info['downside_before_buffer']
            calc_dict['period/time_factor'] = TIME_FACTOR / \
                etf_info['remaining_outcome_period']
            calc_dict['rank'] = calc_dict['risk/reward'] * \
                calc_dict['period/time_factor']

            calc_dict['downside/buffer'] = etf_info['downside_before_buffer'] / \
                etf_info['remaining_buffer']

            calc_dict['score'] = calc_dict['raw_sum_mult'] + \
                calc_dict['rank'] + calc_dict['downside/buffer']

            calc_list.append(calc_dict)

    return calc_list


def sort_calc_list(calc_list, sort_key='score'):
    return sorted(calc_list, key=lambda calc_dict: calc_dict[sort_key], reverse=True)


def as_text(value):
    if value is None or (isinstance(value, str) and '=' in value):
        return ""
    return str(value)


def resize_columns(worksheet):
    for column_cells in worksheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0]
                                    .column_letter].width = length + 2


def format_percentages(worksheet):
    header_row = worksheet[1]

    for cell in header_row:
        # if this should be represented as a % row, then format
        if cell.value and '%' in cell.value:
            for cell in worksheet[cell.column_letter][1:]:
                cell.number_format = '0.00%'


if __name__ == "__main__":
    excel_writer()
