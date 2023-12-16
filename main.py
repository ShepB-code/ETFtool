import json
import openpyxl as xl

def main():
    # open workbook
    wb = xl.load_workbook('first_fund_report.xlsx')
    ws = wb.active
    ws.title = "ETFs"

    ws['A1'] = 'Score Rank'
    ws['B1'] = 'Index'
    ws['C1'] = 'Provider'
    ws['D1'] = 'Ticker'
    ws['E1'] = 'Remaining Cap'
    ws['F1'] = 'Remaining Buffer'
    ws['G1'] = 'Downside Before Buffer'
    ws['H1'] = 'Remaining Outcome Period'

    # get json
    with open('first_trust_outcome_period_values.json', 'r') as json_file: 
        first_trust_json_data = json.load(json_file)
        
        for i, etf_data in enumerate(first_trust_json_data.values()):
            ws.cell(i + 2, 2).value = i + 1
            ws.cell(i + 2, 3).value = "First Trust"
            ws.cell(i + 2, 4).value = etf_data['ticker']
            ws.cell(i + 2, 5).value = etf_data['remaining_cap']
            ws.cell(i + 2, 6).value = etf_data['remaining_buffer']
            ws.cell(i + 2, 7).value = etf_data['downside_before_buffer']
            ws.cell(i + 2, 8).value = etf_data['remaining_outcome_period']

    wb.save("first_fund_report.xlsx")

if __name__ == "__main__":
    main()